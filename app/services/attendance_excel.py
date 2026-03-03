from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.styles.borders import Side
from openpyxl.utils import get_column_letter


# ---- status definitions (frontend attendance.tsx compatible) ----
ATTENDANCE_STATUSES = {
    "出席": {"counted": True, "score": 1.0},
    "欠席": {"counted": True, "score": 0.0},
    "遅刻": {"counted": True, "score": 0.5},
    "早退": {"counted": True, "score": 0.5},
    "講習": {"counted": False, "score": 1.0},
    "無欠": {"counted": True, "score": 0.0},
}

STATUS_FILL = {
    "出席": PatternFill("solid", fgColor="c7f9d9"),
    "欠席": PatternFill("solid", fgColor="ffc9c9"),
    "遅刻": PatternFill("solid", fgColor="fee685"),
    "早退": PatternFill("solid", fgColor="ffd6a7"),
    "講習": PatternFill("solid", fgColor="bedbff"),
    "": PatternFill("solid", fgColor="FFFFFF"),
}

MUKETSU_FILL = PatternFill(patternType="darkTrellis", fgColor="9B9162", bgColor="646464")

HEADER_LEFT_FILL = PatternFill("solid", fgColor="3B82F6")
HEADER_MONTH_FILL = PatternFill("solid", fgColor="E0E7FF")
HEADER_DAY_FILL = PatternFill("solid", fgColor="F9FAFB")
PART_ROW_FILL = PatternFill("solid", fgColor="EEF2FF")

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


@dataclass(frozen=True)
class MemberLite:
    id: str
    part_value: str
    part_en_short: str
    generation: int
    name: str
    name_kana: str


def month_key(d: dt.date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def calc_rate_from_attendances(attendances: Iterable[str]) -> Optional[float]:
    total = 0
    score = 0.0
    for a in attendances:
        st = ATTENDANCE_STATUSES.get(a)
        if not st:
            continue
        if not st["counted"]:
            continue
        total += 1
        score += float(st["score"])
    if total == 0:
        return None
    return round(score / total * 100.0, 1)


def apply_cell_base(cell):
    cell.border = BORDER
    cell.font = Font(name="メイリオ", size=10, bold=False)
    cell.alignment = CENTER


def write_rate(cell, rate: Optional[float], *, base_fill: Optional[PatternFill] = None, bold: bool = False):
    apply_cell_base(cell)
    if base_fill is not None:
        cell.fill = base_fill
    cell.alignment = RIGHT
    if rate is None:
        cell.value = None
        cell.number_format = "0.0"
        cell.font = Font(name="メイリオ", size=10, bold=bold, color="1F2937")
        return

    # %記号を出さず、85.3 のような 0〜100 の数値として表示する
    cell.value = float(rate)
    # 100%相当のときだけ「100」にして場所を取らないようにする
    cell.number_format = "0" if abs(float(rate) - 100.0) < 1e-9 else "0.0"

    if rate < 80:
        cell.font = Font(name="メイリオ", size=10, bold=True, color="DC2626")
    else:
        cell.font = Font(name="メイリオ", size=10, bold=bold, color="111827")


def _fetch_part_en_short_map(settings_yml_path: Path) -> dict[str, str]:
    try:
        import yaml  # type: ignore

        data = yaml.safe_load(settings_yml_path.read_text(encoding="utf-8"))
        parts = data.get("parts") if isinstance(data, dict) else None
        if isinstance(parts, dict):
            out: dict[str, str] = {}
            for k, v in parts.items():
                if isinstance(v, dict) and "enShort" in v:
                    out[str(k)] = str(v["enShort"])
            return out
    except Exception:
        return {}
    return {}


def build_attendance_xlsx_bytes(
    *,
    schedules: list[dt.date],
    members: list[MemberLite],
    actual_map: dict[tuple[str, dt.date], str],
    pre_map: dict[tuple[str, dt.date], str],
    months: Optional[list[str]] = None,
    expand_months: Optional[set[str]] = None,
    both_sheets: bool = True,
    display_mode: str = "actual",
    day_col_width: float = 5,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if both_sheets:
        _export_sheet(
            wb,
            title="実際",
            display_mode="actual",
            schedules=schedules,
            members=members,
            actual_map=actual_map,
            pre_map=pre_map,
            months=months,
            expand_months=expand_months,
            day_col_width=day_col_width,
        )
        _export_sheet(
            wb,
            title="事前",
            display_mode="pre",
            schedules=schedules,
            members=members,
            actual_map=actual_map,
            pre_map=pre_map,
            months=months,
            expand_months=expand_months,
            day_col_width=day_col_width,
        )
    else:
        _export_sheet(
            wb,
            title=("事前" if display_mode == "pre" else "実際"),
            display_mode=display_mode,
            schedules=schedules,
            members=members,
            actual_map=actual_map,
            pre_map=pre_map,
            months=months,
            expand_months=expand_months,
            day_col_width=day_col_width,
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _export_sheet(
    wb: Workbook,
    *,
    title: str,
    display_mode: str,
    schedules: list[dt.date],
    members: list[MemberLite],
    actual_map: dict[tuple[str, dt.date], str],
    pre_map: dict[tuple[str, dt.date], str],
    months: Optional[list[str]],
    expand_months: Optional[set[str]],
    day_col_width: float,
) -> None:
    ws = wb.create_sheet(title=title)

    schedule_months: list[str] = []
    for d in schedules:
        mk = month_key(d)
        if mk not in schedule_months:
            schedule_months.append(mk)
    schedule_months.sort()

    month_order = [m for m in months if m in schedule_months] if months else schedule_months

    schedule_days_by_month: dict[str, list[int]] = {m: [] for m in month_order}
    for d in schedules:
        mk = month_key(d)
        if mk in schedule_days_by_month:
            schedule_days_by_month[mk].append(d.day)
    for m in schedule_days_by_month:
        schedule_days_by_month[m] = sorted(set(schedule_days_by_month[m]))

    members_sorted = sorted(members, key=lambda m: (m.part_value, m.generation, m.name_kana))
    by_part: dict[str, list[MemberLite]] = {}
    for m in members_sorted:
        by_part.setdefault(m.part_value, []).append(m)

    expanded = set(month_order) if expand_months is None else set(expand_months)

    def get_current_attendance(member_id: str, date: dt.date) -> str:
        if display_mode == "pre":
            return pre_map.get((member_id, date), "")
        return actual_map.get((member_id, date), "")

    def has_difference(member_id: str, date: dt.date) -> bool:
        if date > dt.date.today():
            return False
        a = actual_map.get((member_id, date))
        p = pre_map.get((member_id, date))
        if not a or not p:
            return False
        return a != p

    def diff_tooltip(member_id: str, date: dt.date) -> str:
        a = actual_map.get((member_id, date), "-")
        p = pre_map.get((member_id, date), "-")
        return f"確定: {a}" if display_mode == "pre" else f"事前: {p}"

    columns: list[tuple[str, Optional[int]]] = []
    for mk in month_order:
        columns.append((mk, None))
        if mk in expanded:
            for day in schedule_days_by_month.get(mk, []):
                columns.append((mk, day))

    # ヘッダは2行使う（1行目: 見出し、2行目: 出席率）
    header_row = 1
    header_rate_row = header_row + 1
    ws.row_dimensions[header_row].height = 48
    ws.row_dimensions[header_rate_row].height = 18

    tl = ws.cell(row=header_row, column=1, value="全体")
    apply_cell_base(tl)
    tl.fill = HEADER_LEFT_FILL
    tl.font = Font(name="メイリオ", bold=True, color="FFFFFF", size=16)
    tl.alignment = CENTER

    # 「全体」は2行分を縦結合しておく
    ws.merge_cells(start_row=header_row, start_column=1, end_row=header_rate_row, end_column=1)

    ws.column_dimensions["A"].width = 22

    col_idx = 2
    idx_month_rate_col: dict[str, int] = {}
    idx_day_col: dict[tuple[str, int], int] = {}

    for mk, day in columns:
        if day is None:
            idx_month_rate_col[mk] = col_idx
        else:
            idx_day_col[(mk, day)] = col_idx

        # 1行目セル（見出し）
        c = ws.cell(row=header_row, column=col_idx)
        apply_cell_base(c)

        if day is None:
            ws.column_dimensions[get_column_letter(col_idx)].width = 7
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = day_col_width

        if day is None:
            # ---- 月セル: 1行目=◯月, 2行目=月別出席率（write_rateでスタイル統一） ----
            month_num = int(mk.split("-")[1])

            # 1行目（◯月）
            c.value = f"{month_num}月"
            c.fill = HEADER_MONTH_FILL
            c.font = Font(name="メイリオ", bold=True, color="1D4ED8", size=12)
            c.alignment = CENTER

            # 2行目（月別出席率）
            attendances: list[str] = []
            for part_mems in by_part.values():
                for mem in part_mems:
                    for d in schedule_days_by_month.get(mk, []):
                        date = dt.date.fromisoformat(f"{mk}-{d:02d}")
                        a = actual_map.get((mem.id, date))
                        if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                            attendances.append(a)
            rate = calc_rate_from_attendances(attendances)

            rate_cell = ws.cell(row=header_rate_row, column=col_idx)
            write_rate(rate_cell, rate, base_fill=HEADER_MONTH_FILL, bold=True)

            # 80%未満のときは月(上段)も赤に寄せて見た目を揃える
            if rate is not None and rate < 80:
                c.font = Font(name="メイリオ", bold=True, color="DC2626", size=12)

        else:
            # ---- 日セル: 1行目=日付, 2行目=日別出席率（write_rateでスタイル統一） ----
            date = dt.date.fromisoformat(f"{mk}-{day:02d}")
            day_att: list[str] = []
            for part_mems in by_part.values():
                for mem in part_mems:
                    a = actual_map.get((mem.id, date))
                    if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                        day_att.append(a)
            day_rate = calc_rate_from_attendances(day_att)

            # 1行目（日付）
            c.value = f"{day}"
            c.fill = HEADER_DAY_FILL
            c.font = Font(name="メイリオ", bold=True, color="111827", size=12)
            c.alignment = CENTER

            # 2行目（日別出席率）
            day_rate_cell = ws.cell(row=header_rate_row, column=col_idx)
            write_rate(day_rate_cell, day_rate, base_fill=HEADER_DAY_FILL, bold=True)

            # 80%未満のときは日付(上段)も赤に寄せて見た目を揃える
            if day_rate is not None and day_rate < 80:
                c.font = Font(name="メイリオ", bold=True, color="DC2626", size=12)


        col_idx += 1

    max_col = col_idx - 1
    ws.freeze_panes = "B3"

    row = header_rate_row + 1
    # enShort map from settings.yml (optional)
    en_short_map = _fetch_part_en_short_map(Path("settings.yml"))

    for part_value in sorted(by_part.keys()):
        mems = by_part[part_value]
        display_part = en_short_map.get(part_value, mems[0].part_en_short if mems else part_value)

        part_cell = ws.cell(row=row, column=1, value=display_part)
        apply_cell_base(part_cell)
        part_cell.fill = PART_ROW_FILL
        part_cell.font = Font(name="メイリオ", bold=True, color="1D4ED8", size=16)
        part_cell.alignment = LEFT

        for mk, day in columns:
            if day is None:
                cell = ws.cell(row=row, column=idx_month_rate_col[mk])
                vals: list[str] = []
                for mem in mems:
                    for d in schedule_days_by_month.get(mk, []):
                        date = dt.date.fromisoformat(f"{mk}-{d:02d}")
                        a = actual_map.get((mem.id, date))
                        if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                            vals.append(a)
                write_rate(cell, calc_rate_from_attendances(vals), base_fill=PART_ROW_FILL, bold=False)
            else:
                cell = ws.cell(row=row, column=idx_day_col[(mk, day)])
                date = dt.date.fromisoformat(f"{mk}-{day:02d}")
                vals: list[str] = []
                for mem in mems:
                    a = actual_map.get((mem.id, date))
                    if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                        vals.append(a)
                write_rate(cell, calc_rate_from_attendances(vals), base_fill=PART_ROW_FILL, bold=False)
                cell.alignment = RIGHT

        row += 1

        for mem in mems:
            name_cell = ws.cell(row=row, column=1, value=mem.name)
            apply_cell_base(name_cell)
            name_cell.alignment = LEFT
            name_cell.fill = PatternFill("solid", fgColor="FFFFFF")
            name_cell.font = Font(name="メイリオ", bold=False, color="111827", size=11)

            for mk, day in columns:
                if day is None:
                    cell = ws.cell(row=row, column=idx_month_rate_col[mk])
                    vals: list[str] = []
                    for d in schedule_days_by_month.get(mk, []):
                        date = dt.date.fromisoformat(f"{mk}-{d:02d}")
                        a = actual_map.get((mem.id, date))
                        if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                            vals.append(a)
                    rate = calc_rate_from_attendances(vals)
                    write_rate(cell, rate, base_fill=None, bold=False)
                else:
                    date = dt.date.fromisoformat(f"{mk}-{day:02d}")
                    val = get_current_attendance(mem.id, date)
                    shown = (val[:3] if val else "-")
                    cell = ws.cell(row=row, column=idx_day_col[(mk, day)], value=shown)
                    apply_cell_base(cell)
                    cell.alignment = CENTER

                    if val == "無欠":
                        cell.fill = MUKETSU_FILL
                        cell.font = Font(name="メイリオ", bold=False, color="FFFFFF", size=10)
                    else:
                        cell.fill = STATUS_FILL.get(val, STATUS_FILL[""])
                        cell.font = Font(name="メイリオ", bold=False, color="111827", size=(9 if len(shown) > 2 else 10))

                    if has_difference(mem.id, date):
                        cell.comment = Comment(diff_tooltip(mem.id, date), "diff")

            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{row-1}"
