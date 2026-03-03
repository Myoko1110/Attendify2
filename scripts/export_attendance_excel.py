"""Export attendance table to Excel (React attendance.tsx compatible).

React `src/sections/attendance-n/attendance.tsx` の表にできるだけ一致する形で、
SQLite(`attendify.db`)のデータから出欠表をExcel(xlsx)として出力します。

再現ポイント:
- ヘッダ:
  - 左上: 「全体」
  - 月ごとに、(月率セル) + (展開時のみ: 日セル×スケジュール日数)
  - 日セルには日付(1-31) + 日別出席率(全体)
  - 月率セルには「◯月」+ 月別出席率(全体)
- ボディ:
  - パート行: part.enShort を表示し、(月率 + 日別率)
  - 部員行: 名前、(月率 + 出欠)
- 出席率:
  - 80%未満は赤字
  - 「講習」は counted=false のため出席率計算から除外
- 出欠セル:
  - 文字は value.substring(0, 3) 相当（最大3文字）
  - 色は概ねフロントの指定に合わせた近似色
  - 無欠は斜線パターン(近似)
- 差分:
  - 今日以前で pre と actual が違う場合、左上に赤三角 + コメント(ツールチップ代替)

Usage (PowerShell):
  python scripts/export_attendance_excel.py --database attendify.db --out attendance.xlsx

Options:
  --display-mode actual|pre
  --months 2026-03 2026-04  (指定しない場合はスケジュールにある月すべて)
  --expand-months 2026-03  (指定月だけ展開。未指定は全て展開)
  --grades 58 59           (generation フィルタ)
  --group-ids <uuid> ...   (グループフィルタ: member_groups)

"""

from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.styles.borders import Side
from openpyxl.utils import get_column_letter


# ---- attendance.tsx compatible statuses ----
# counted=false: 講習
ATTENDANCE_STATUSES = {
    "出席": {"counted": True, "score": 1.0},
    "欠席": {"counted": True, "score": 0.0},
    "遅刻": {"counted": True, "score": 0.5},
    "早退": {"counted": True, "score": 0.5},
    "講習": {"counted": False, "score": 1.0},
    "無欠": {"counted": True, "score": 0.0},
}

# フロントの見た目に寄せた近似色
STATUS_FILL = {
    "出席": PatternFill("solid", fgColor="C6EFCE"),  # green-200
    "欠席": PatternFill("solid", fgColor="FFC7CE"),  # red-200
    "遅刻": PatternFill("solid", fgColor="FCE4D6"),  # orange-200
    "早退": PatternFill("solid", fgColor="FFE699"),  # amber-200
    "講習": PatternFill("solid", fgColor="BDD7EE"),  # blue-200
    "": PatternFill("solid", fgColor="FFFFFF"),
}

# 無欠は「斜線」(repeating-linear-gradient) をExcelのpatternで近似
MUKETSU_FILL = PatternFill(patternType="darkTrellis", fgColor="9B9162", bgColor="646464")

HEADER_LEFT_FILL = PatternFill("solid", fgColor="3B82F6")  # blue-500
HEADER_MONTH_FILL = PatternFill("solid", fgColor="E0E7FF")  # indigo-100
HEADER_DAY_FILL = PatternFill("solid", fgColor="F9FAFB")  # gray-50
PART_ROW_FILL = PatternFill("solid", fgColor="EEF2FF")  # indigo-50

THIN = Side(style="thin", color="D1D5DB")  # gray-300
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


@dataclass(frozen=True)
class GroupRow:
    id: str
    display_name: str


@dataclass(frozen=True)
class MemberRow:
    id: str
    part_value: str
    part_en_short: str
    generation: int
    name: str
    name_kana: str


@dataclass(frozen=True)
class ScheduleRow:
    date: dt.date


def month_key(d: dt.date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def sqlite_connect(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(path))
    conn.row_factory = sqlite3.Row
    return conn


def fetch_member_group_ids(conn: sqlite3.Connection) -> dict[str, set[str]]:
    rows = conn.execute("SELECT member_id, group_id FROM member_groups").fetchall()
    out: dict[str, set[str]] = {}
    for r in rows:
        out.setdefault(str(r["member_id"]), set()).add(str(r["group_id"]))
    return out


def fetch_members(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT id, part, generation, name, name_kana FROM members"
    ).fetchall()


def fetch_schedules(conn: sqlite3.Connection) -> list[ScheduleRow]:
    rows = conn.execute("SELECT date FROM schedules ORDER BY date").fetchall()
    return [ScheduleRow(date=dt.date.fromisoformat(r["date"])) for r in rows]


def fetch_attendance_map(conn: sqlite3.Connection) -> dict[tuple[str, dt.date], str]:
    rows = conn.execute("SELECT member_id, date, attendance FROM attendances").fetchall()
    out: dict[tuple[str, dt.date], str] = {}
    for r in rows:
        out[(str(r["member_id"]), dt.date.fromisoformat(r["date"]))] = str(r["attendance"])
    return out


def fetch_pre_attendance_map(conn: sqlite3.Connection) -> dict[tuple[str, dt.date], str]:
    rows = conn.execute("SELECT member_id, date, attendance FROM pre_attendances").fetchall()
    out: dict[tuple[str, dt.date], str] = {}
    for r in rows:
        if r["member_id"] is None:
            continue
        out[(str(r["member_id"]), dt.date.fromisoformat(r["date"]))] = str(r["attendance"])
    return out


def fetch_part_en_short(settings_yml_path: Path) -> dict[str, str]:
    # backend settings.yml に Part 表示名があれば使う。無い場合はvalueをそのまま表示。
    # このリポジトリのsettings.ymlは環境によって異なる可能性があるので、落ちても運用可にする。
    try:
        import yaml  # type: ignore

        data = yaml.safe_load(settings_yml_path.read_text(encoding="utf-8"))
        # 想定: settings.yml に part 表示設定がある場合
        # 例: parts: { flute: { enShort: 'Fl' }, ... }
        parts = data.get("parts") if isinstance(data, dict) else None
        if isinstance(parts, dict):
            out = {}
            for k, v in parts.items():
                if isinstance(v, dict) and "enShort" in v:
                    out[str(k)] = str(v["enShort"])
            if out:
                return out
    except Exception:
        pass
    return {}


def calc_rate_from_attendances(attendances: Iterable[str]) -> Optional[float]:
    total = 0
    score = 0.0
    for a in attendances:
        st = ATTENDANCE_STATUSES.get(a)
        if not st:
            continue
        if not st["counted"]:
            continue
        total += 1
        score += float(st["score"])
    if total == 0:
        return None
    # React側 Attendances.calcRate() は backend同様 0.1刻みと推測されるため、ここも0.1刻みに揃える
    return round(score / total * 100.0, 1)


def apply_cell_base(cell):
    cell.border = BORDER
    cell.font = Font(name="メイリオ", size=10, bold=False)
    cell.alignment = CENTER


def write_rate(cell, rate: Optional[float], *, base_fill: Optional[PatternFill] = None, bold: bool = False):
    apply_cell_base(cell)
    if base_fill is not None:
        cell.fill = base_fill
    cell.alignment = RIGHT
    if rate is None:
        cell.value = None
        cell.number_format = "0.0%"
        cell.font = Font(name="メイリオ", size=10, bold=bold, color="1F2937")  # gray-800
        return

    cell.value = rate / 100.0
    cell.number_format = "0.0%"

    if rate < 80:
        cell.font = Font(name="メイリオ", size=10, bold=True, color="DC2626")  # red-600
    else:
        cell.font = Font(name="メイリオ", size=10, bold=bold, color="2563EB" if bold else "111827")  # blue-600-ish


# NOTE:
# Reactの「左上赤三角」はCSSで擬似要素を描画していますが、Excelではセルの"コメント"を付けると
# 左上に三角が表示されるため、ここではコメントで再現します（互換性と安定性優先）。


def export_sheet(
    wb: Workbook,
    *,
    title: str,
    db_path: Path,
    display_mode: str,
    months: Optional[list[str]],
    expand_months: Optional[set[str]],
    grades: Optional[set[int]],
    group_ids: Optional[set[str]],
) -> None:
    """1つのシートに出欠表を書き込む。"""
    conn = sqlite_connect(db_path)
    try:
        member_group_ids = fetch_member_group_ids(conn)
        member_rows = fetch_members(conn)
        schedules = fetch_schedules(conn)
        actual_map = fetch_attendance_map(conn)
        pre_map = fetch_pre_attendance_map(conn)
    finally:
        conn.close()

    # ---- workbook/sheet ----
    ws = wb.create_sheet(title=title)

    # 月一覧（React: scheduleから抽出）
    schedule_months: list[str] = []
    for s in schedules:
        mk = month_key(s.date)
        if mk not in schedule_months:
            schedule_months.append(mk)
    schedule_months.sort()

    if months:
        month_order = [m for m in months if m in schedule_months]
    else:
        month_order = schedule_months

    # 月→スケジュール日
    schedule_days_by_month: dict[str, list[int]] = {m: [] for m in month_order}
    for s in schedules:
        mk = month_key(s.date)
        if mk in schedule_days_by_month:
            schedule_days_by_month[mk].append(s.date.day)
    for m in schedule_days_by_month:
        schedule_days_by_month[m] = sorted(set(schedule_days_by_month[m]))

    # settings.yml から enShort が取れれば使う
    part_en_short_map = fetch_part_en_short(Path("settings.yml"))

    # Memberフィルタ & part情報整形（React: advisor除外、part/generation/nameKanaソート、group filter）
    members: list[MemberRow] = []
    for r in member_rows:
        mid = str(r["id"])
        part_value = str(r["part"])

        if part_value.lower() == "advisor":
            continue

        gen = int(r["generation"])
        if grades is not None and gen not in grades:
            continue

        if group_ids is not None and len(group_ids) > 0:
            mg = member_group_ids.get(mid, set())
            if not any(g in mg for g in group_ids):
                continue

        en_short = part_en_short_map.get(part_value, part_value)

        members.append(
            MemberRow(
                id=mid,
                part_value=part_value,
                part_en_short=en_short,
                generation=gen,
                name=str(r["name"]),
                name_kana=str(r["name_kana"]),
            )
        )

    members.sort(key=lambda m: (m.part_value, m.generation, m.name_kana))

    # partごとにグループ化
    by_part: dict[str, list[MemberRow]] = {}
    for m in members:
        by_part.setdefault(m.part_value, []).append(m)

    # 展開月
    if expand_months is None:
        expanded = set(month_order)  # デフォルト: 全展開
    else:
        expanded = set(expand_months)

    def get_current_attendance(member_id: str, date: dt.date) -> str:
        if display_mode == "pre":
            return pre_map.get((member_id, date), "")
        return actual_map.get((member_id, date), "")

    def has_difference(member_id: str, date: dt.date) -> bool:
        if date > dt.date.today():
            return False
        a = actual_map.get((member_id, date))
        p = pre_map.get((member_id, date))
        if not a or not p:
            return False
        return a != p

    def diff_tooltip(member_id: str, date: dt.date) -> str:
        a = actual_map.get((member_id, date), "-")
        p = pre_map.get((member_id, date), "-")
        return f"確定: {a}" if display_mode == "pre" else f"事前: {p}"

    ws.title = title

    # 列構築（Reactの月折りたたみ構造に合わせる）
    columns: list[tuple[str, Optional[int]]] = []
    for mk in month_order:
        columns.append((mk, None))
        if mk in expanded:
            for day in schedule_days_by_month.get(mk, []):
                columns.append((mk, day))

    header_row = 1
    ws.row_dimensions[header_row].height = 36

    tl = ws.cell(row=header_row, column=1, value="全体")
    apply_cell_base(tl)
    tl.fill = HEADER_LEFT_FILL
    tl.font = Font(name="メイリオ", bold=True, color="FFFFFF", size=16)
    tl.alignment = CENTER

    ws.column_dimensions["A"].width = 22

    col_idx = 2
    idx_month_rate_col: dict[str, int] = {}
    idx_day_col: dict[tuple[str, int], int] = {}

    for mk, day in columns:
        if day is None:
            idx_month_rate_col[mk] = col_idx
        else:
            idx_day_col[(mk, day)] = col_idx

        c = ws.cell(row=header_row, column=col_idx)
        apply_cell_base(c)
        c.border = BORDER

        if day is None:
            ws.column_dimensions[get_column_letter(col_idx)].width = 7
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 7

        if day is None:
            month_num = int(mk.split("-")[1])
            attendances: list[str] = []
            for part_mems in by_part.values():
                for mem in part_mems:
                    for d in schedule_days_by_month.get(mk, []):
                        date = dt.date.fromisoformat(f"{mk}-{d:02d}")
                        a = actual_map.get((mem.id, date))
                        if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                            attendances.append(a)
            rate = calc_rate_from_attendances(attendances)

            c.value = f"{month_num}月\n{rate}%" if rate is not None else f"{month_num}月"
            c.fill = HEADER_MONTH_FILL
            c.font = Font(name="メイリオ", bold=True, color="1D4ED8", size=12)
            c.alignment = CENTER_WRAP
            if rate is not None and rate < 80:
                c.font = Font(name="メイリオ", bold=True, color="DC2626", size=12)
        else:
            date = dt.date.fromisoformat(f"{mk}-{day:02d}")
            day_att: list[str] = []
            for part_mems in by_part.values():
                for mem in part_mems:
                    a = actual_map.get((mem.id, date))
                    if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                        day_att.append(a)
            day_rate = calc_rate_from_attendances(day_att)
            c.value = f"{day}\n{day_rate}%" if day_rate is not None else f"{day}"
            c.fill = HEADER_DAY_FILL
            c.font = Font(name="メイリオ", bold=True, color="111827", size=12)
            c.alignment = CENTER_WRAP
            if day_rate is not None and day_rate < 80:
                c.font = Font(name="メイリオ", bold=True, color="DC2626", size=12)

        col_idx += 1

    max_col = col_idx - 1
    ws.freeze_panes = "B2"

    row = 2
    for part_value in sorted(by_part.keys()):
        mems = by_part[part_value]

        part_cell = ws.cell(row=row, column=1, value=mems[0].part_en_short if mems else part_value)
        apply_cell_base(part_cell)
        part_cell.fill = PART_ROW_FILL
        part_cell.font = Font(name="メイリオ", bold=True, color="1D4ED8", size=16)
        part_cell.alignment = LEFT

        for mk, day in columns:
            if day is None:
                cell = ws.cell(row=row, column=idx_month_rate_col[mk])
                vals: list[str] = []
                for mem in mems:
                    for d in schedule_days_by_month.get(mk, []):
                        date = dt.date.fromisoformat(f"{mk}-{d:02d}")
                        a = actual_map.get((mem.id, date))
                        if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                            vals.append(a)
                write_rate(cell, calc_rate_from_attendances(vals), base_fill=PART_ROW_FILL, bold=False)
            else:
                cell = ws.cell(row=row, column=idx_day_col[(mk, day)])
                date = dt.date.fromisoformat(f"{mk}-{day:02d}")
                vals: list[str] = []
                for mem in mems:
                    a = actual_map.get((mem.id, date))
                    if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                        vals.append(a)
                write_rate(cell, calc_rate_from_attendances(vals), base_fill=PART_ROW_FILL, bold=False)
                cell.alignment = RIGHT

        row += 1

        for mem in mems:
            name_cell = ws.cell(row=row, column=1, value=mem.name)
            apply_cell_base(name_cell)
            name_cell.alignment = LEFT
            name_cell.fill = PatternFill("solid", fgColor="FFFFFF")
            name_cell.font = Font(name="メイリオ", bold=False, color="111827", size=11)

            for mk, day in columns:
                if day is None:
                    cell = ws.cell(row=row, column=idx_month_rate_col[mk])
                    vals: list[str] = []
                    for d in schedule_days_by_month.get(mk, []):
                        date = dt.date.fromisoformat(f"{mk}-{d:02d}")
                        a = actual_map.get((mem.id, date))
                        if a and ATTENDANCE_STATUSES.get(a, {}).get("counted"):
                            vals.append(a)
                    rate = calc_rate_from_attendances(vals)
                    write_rate(cell, rate, base_fill=None, bold=False)
                else:
                    date = dt.date.fromisoformat(f"{mk}-{day:02d}")
                    val = get_current_attendance(mem.id, date)
                    shown = (val[:3] if val else "-")
                    cell = ws.cell(row=row, column=idx_day_col[(mk, day)], value=shown)
                    apply_cell_base(cell)
                    cell.alignment = CENTER

                    if val == "無欠":
                        cell.fill = MUKETSU_FILL
                        cell.font = Font(name="メイリオ", bold=False, color="FFFFFF", size=10)
                    else:
                        cell.fill = STATUS_FILL.get(val, STATUS_FILL[""])
                        cell.font = Font(name="メイリオ", bold=False, color="111827", size=(9 if len(shown) > 2 else 10))

                    if has_difference(mem.id, date):
                        cell.comment = Comment(diff_tooltip(mem.id, date), "diff")

            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{row-1}"


def export_excel(
    *,
    db_path: Path,
    out_path: Path,
    display_mode: str,
    months: Optional[list[str]],
    expand_months: Optional[set[str]],
    grades: Optional[set[int]],
    group_ids: Optional[set[str]],
    both_sheets: bool,
) -> None:
    wb = Workbook()
    # デフォルトシート削除（create_sheetで作り直す）
    default_ws = wb.active
    wb.remove(default_ws)

    if both_sheets:
        export_sheet(
            wb,
            title="実際",
            db_path=db_path,
            display_mode="actual",
            months=months,
            expand_months=expand_months,
            grades=grades,
            group_ids=group_ids,
        )
        export_sheet(
            wb,
            title="事前",
            db_path=db_path,
            display_mode="pre",
            months=months,
            expand_months=expand_months,
            grades=grades,
            group_ids=group_ids,
        )
    else:
        export_sheet(
            wb,
            title=("事前" if display_mode == "pre" else "実際"),
            db_path=db_path,
            display_mode=display_mode,
            months=months,
            expand_months=expand_months,
            grades=grades,
            group_ids=group_ids,
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--database", default="attendify.db", help="SQLiteファイルパス")
    p.add_argument("--out", default="attendance.xlsx", help="出力xlsxパス")
    p.add_argument("--display-mode", choices=["actual", "pre"], default="actual")
    p.add_argument("--months", nargs="+", help="対象月(YYYY-MM)。未指定なら全月")
    p.add_argument("--expand-months", nargs="+", help="展開する月(YYYY-MM)。未指定なら全月展開")
    p.add_argument("--grades", nargs="+", type=int, help="generation（学年）フィルタ")
    p.add_argument("--group-ids", nargs="+", help="グループIDフィルタ")
    p.add_argument("--both-sheets", action="store_true", help="1つのxlsxに確定出欠と事前出欠の2シートを出力")

    args = p.parse_args()

    export_excel(
        db_path=Path(args.database),
        out_path=Path(args.out),
        display_mode=args.display_mode,
        months=args.months,
        expand_months=set(args.expand_months) if args.expand_months else None,
        grades=set(args.grades) if args.grades else None,
        group_ids=set(args.group_ids) if args.group_ids else None,
        both_sheets=bool(args.both_sheets),
    )


if __name__ == "__main__":
    main()
